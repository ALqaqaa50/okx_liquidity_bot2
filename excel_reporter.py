"""
Excel Reporter Module
Exports trading bot data to Excel format
"""

import json
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger("ExcelReporter")


class ExcelReporter:
    """Generate Excel reports from trading bot data"""
    
    def __init__(self, data_dir: Path = Path("./data")):
        self.data_dir = data_dir
    
    def _load_jsonl(self, filename: str) -> List[Dict[str, Any]]:
        """Load JSONL file and return list of dictionaries"""
        filepath = self.data_dir / filename
        if not filepath.exists():
            return []
        
        data = []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        data.append(json.loads(line))
        except Exception as e:
            logger.error(f"Error loading {filename}: {e}")
        
        return data
    
    def create_comprehensive_report(self, output_path: str = None) -> str:
        """
        Create a comprehensive Excel report with multiple sheets
        Returns the path to the generated file
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.data_dir / f"trading_report_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        
        # Load all data
        snapshots = self._load_jsonl("market_snapshots.jsonl")
        metrics = self._load_jsonl("metrics.jsonl")
        orders = self._load_jsonl("orders.jsonl")
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Market Snapshots
            if snapshots:
                df_snapshots = pd.DataFrame(snapshots)
                # Convert timestamp to readable format
                if 'ts' in df_snapshots.columns:
                    df_snapshots['datetime'] = pd.to_datetime(df_snapshots['ts'], unit='s')
                df_snapshots.to_excel(writer, sheet_name='Market Snapshots', index=False)
            
            # Sheet 2: Orders
            if orders:
                df_orders = pd.DataFrame(orders)
                if 'ts' in df_orders.columns:
                    df_orders['datetime'] = pd.to_datetime(df_orders['ts'], unit='s')
                df_orders.to_excel(writer, sheet_name='Orders', index=False)
            
            # Sheet 3: API Metrics
            if metrics:
                df_metrics = pd.DataFrame(metrics)
                if 'ts' in df_metrics.columns:
                    df_metrics['datetime'] = pd.to_datetime(df_metrics['ts'], unit='s')
                df_metrics.to_excel(writer, sheet_name='API Metrics', index=False)
            
            # Sheet 4: Summary Statistics
            if snapshots:
                summary_data = self._generate_summary_stats(snapshots, orders)
                df_summary = pd.DataFrame([summary_data])
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Sheet 5: Performance Analysis
            if orders:
                perf_data = self._analyze_performance(orders)
                if perf_data:
                    df_perf = pd.DataFrame(perf_data)
                    df_perf.to_excel(writer, sheet_name='Performance', index=False)
        
        # Apply formatting
        self._format_workbook(output_path)
        
        logger.info(f"Excel report generated: {output_path}")
        return str(output_path)
    
    def _generate_summary_stats(self, snapshots: List[Dict], orders: List[Dict]) -> Dict:
        """Generate summary statistics"""
        summary = {
            'Report Generated': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'Total Snapshots': len(snapshots),
            'Total Orders': len(orders),
        }
        
        if snapshots:
            df = pd.DataFrame(snapshots)
            if 'last_price' in df.columns:
                summary['Avg Price'] = df['last_price'].mean()
                summary['Min Price'] = df['last_price'].min()
                summary['Max Price'] = df['last_price'].max()
            
            if 'decision' in df.columns:
                summary['Long Signals'] = (df['decision'] == 'LONG').sum()
                summary['Short Signals'] = (df['decision'] == 'SHORT').sum()
                summary['Flat Signals'] = (df['decision'] == 'FLAT').sum()
        
        return summary
    
    def _analyze_performance(self, orders: List[Dict]) -> List[Dict]:
        """Analyze trading performance from orders"""
        performance = []
        
        # Simple performance analysis
        for order in orders:
            if 'response' in order and order['response']:
                resp = order['response']
                if isinstance(resp, list) and len(resp) > 0:
                    order_data = resp[0]
                    perf = {
                        'timestamp': datetime.fromtimestamp(order['ts']).strftime("%Y-%m-%d %H:%M:%S"),
                        'order_id': order_data.get('ordId', 'N/A'),
                        'side': order.get('request', {}).get('side', 'N/A'),
                        'pos_side': order.get('request', {}).get('posSide', 'N/A'),
                        'size': order.get('request', {}).get('sz', 'N/A'),
                        'status': order_data.get('sCode', 'N/A'),
                    }
                    performance.append(perf)
        
        return performance
    
    def _format_workbook(self, filepath: Path):
        """Apply formatting to the Excel workbook"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header formatting
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
        except Exception as e:
            logger.error(f"Error formatting workbook: {e}")
    
    def create_daily_report(self) -> str:
        """Create a daily report with today's data"""
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = self.data_dir / f"daily_report_{timestamp}.xlsx"
        return self.create_comprehensive_report(str(output_path))
